import requests
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, deque

def fetch_impact_data(object_type, field, depth=3, rel_type=0, direction='downstream'):
    url = f"http://localhost:18080/api/impact"
    params = {
        'objectType': object_type,
        'field': field,
        'depth': depth,
        'relType': rel_type,
        'direction': direction
    }
    response = requests.get(url, params=params)
    response.raise_for_status()
    return response.json()

def build_paths(nodes, edges, root_field_id, max_depth=None, exclude_types=None):
    graph = defaultdict(list)
    edge_types = {}
    
    exclude_types = exclude_types or []
    
    for edge in edges:
        source = edge['source']
        target = edge['target']
        edge_type = edge.get('type', 'unknown')
        if edge_type not in exclude_types:
            graph[source].append(target)
            edge_types[(source, target)] = edge_type
    
    paths = {}
    visited = set()
    queue = deque([(root_field_id, [], 0)])
    
    while queue:
        current, path, depth = queue.popleft()
        if current in visited:
            continue
        visited.add(current)
        
        if current != root_field_id:
            paths[current] = path
        
        if max_depth is not None and depth >= max_depth:
            continue
        
        for neighbor in graph.get(current, []):
            if neighbor not in visited:
                edge_type = edge_types.get((current, neighbor), 'unknown')
                new_path = path + [(current, neighbor, edge_type)]
                queue.append((neighbor, new_path, depth + 1))
    
    return paths

def format_path(path, nodes_dict, show_types=True):
    if not path:
        return ''
    parts = []
    for source, target, edge_type in path:
        if show_types:
            type_label = {'intra': '触发', 'writeBack': '回写', 'view': '视图'}.get(edge_type, edge_type)
            parts.append(f"{source} --[{type_label}]--> {target}")
        else:
            parts.append(f"{source} → {target}")
    return ' → '.join(parts)

def create_excel(data, output_file, max_depth=None, exclude_types=None, show_types=True):
    wb = Workbook()
    wb.remove(wb.active)
    
    nodes = data.get('nodes', [])
    edges = data.get('edges', [])
    
    root_field_id = None
    root_object = None
    root_field = None
    for node in nodes:
        if node.get('object') == 'InvoiceItem' and (node.get('apiName') == 'originAmount' or node.get('field') == 'originAmount'):
            root_field_id = node.get('id')
            root_object = node.get('object')
            root_field = node.get('apiName') or node.get('field')
            break
    
    if not root_field_id:
        root_field_id = nodes[0].get('id') if nodes else None
        if root_field_id:
            root_node = nodes[0]
            root_object = root_node.get('object')
            root_field = root_node.get('apiName') or root_node.get('field')
    
    exclude_types_set = set(exclude_types) if exclude_types else set()
    
    filtered_edges = [e for e in edges if e.get('type', 'unknown') not in exclude_types_set]
    valid_node_ids = set()
    if root_field_id:
        valid_node_ids.add(root_field_id)
    for edge in filtered_edges:
        valid_node_ids.add(edge['source'])
        valid_node_ids.add(edge['target'])
    
    filtered_nodes = [node for node in nodes if node.get('id') in valid_node_ids]
    
    nodes_dict = {node['id']: node for node in filtered_nodes}
    paths = build_paths(filtered_nodes, filtered_edges, root_field_id, max_depth=max_depth, exclude_types=exclude_types) if root_field_id else {}
    
    objects = defaultdict(list)
    for node in filtered_nodes:
        obj_name = node.get('object')
        objects[obj_name].append(node)
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for obj_name in sorted(objects.keys()):
        sheet = wb.create_sheet(title=obj_name[:31])
        fields = objects[obj_name]
        
        row = 1
        sheet.merge_cells(f'A{row}:J{row}')
        sheet[f'A{row}'] = f'对象: {obj_name}'
        sheet[f'A{row}'].font = Font(bold=True, size=14)
        sheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        affected_fields_info = {}
        root_fields = set()
        field_to_node_id = {}
        node_id_to_field_name = {}
        
        for node in fields:
            field_name = node.get('apiName') or node.get('field', '')
            node_id = node.get('id')
            field_to_node_id[field_name] = node_id
            node_id_to_field_name[node_id] = field_name
        
        for edge in filtered_edges:
            edge_type = edge.get('type', 'unknown')
            source_obj = edge['source'].split('.')[0]
            target_obj = edge['target'].split('.')[0]
            source_field = edge['source'].split('.')[1]
            target_node_id = edge['target']
            
            if target_obj == obj_name:
                if target_node_id not in affected_fields_info:
                    affected_fields_info[target_node_id] = {'types': set(), 'paths': []}
                affected_fields_info[target_node_id]['types'].add(edge_type)
            
            if source_obj == obj_name and source_obj == root_object and source_field == root_field:
                root_fields.add(source_field)
        
        if obj_name == root_object and root_field:
            root_fields = {root_field}
        
        headers = ['字段名', '标题', '类型', '业务类型', '字段类型', '影响类型', '影响路径', '触发公式', '表达式', '虚拟表达式']
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row, col_idx, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        row += 1
        
        for field_node in fields:
            field_name = field_node.get('apiName') or field_node.get('field', '')
            node_id = field_node.get('id')
            is_root = field_name in root_fields
            is_affected = node_id in affected_fields_info
            
            field_type = ''
            impact_types = ''
            impact_path = ''
            
            if is_root:
                field_type = '原始字段'
            elif is_affected:
                field_type = '受影响字段'
                info = affected_fields_info[node_id]
                type_labels = []
                for t in info['types']:
                    type_labels.append({'intra': '触发(Intra)', 'writeBack': '回写(WriteBack)', 'view': '视图(View)'}.get(t, t))
                impact_types = ', '.join(sorted(type_labels))
                
                if node_id in paths:
                    path = paths[node_id]
                    impact_path = format_path(path, nodes_dict, show_types=show_types)
            else:
                field_type = '关联字段'
            
            sheet.cell(row, 1, field_name)
            sheet.cell(row, 2, field_node.get('title', ''))
            sheet.cell(row, 3, field_node.get('type', ''))
            sheet.cell(row, 4, field_node.get('bizType', ''))
            sheet.cell(row, 5, field_type)
            sheet.cell(row, 6, impact_types)
            sheet.cell(row, 7, impact_path)
            sheet.cell(row, 8, field_node.get('triggerExpr', '') or '')
            sheet.cell(row, 9, field_node.get('expression', '') or '')
            sheet.cell(row, 10, field_node.get('virtualExpr', '') or '')
            
            for col_idx in range(1, 11):
                cell = sheet.cell(row, col_idx)
                cell.border = border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
                if is_root:
                    cell.fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
                elif is_affected:
                    cell.fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            
            row += 1
        
        for col_idx in range(1, 11):
            col_letter = get_column_letter(col_idx)
            if col_idx == 1:
                sheet.column_dimensions[col_letter].width = 25
            elif col_idx == 2:
                sheet.column_dimensions[col_letter].width = 20
            elif col_idx in [3, 4, 5, 6]:
                sheet.column_dimensions[col_letter].width = 15
            elif col_idx == 7:
                sheet.column_dimensions[col_letter].width = 60
            else:
                sheet.column_dimensions[col_letter].width = 50
        
        sheet.row_dimensions[1].height = 30
        sheet.row_dimensions[2].height = 25
    
    wb.save(output_file)
    print(f"Excel 文件已保存: {output_file}")

if __name__ == '__main__':
    import sys
    
    object_type = 'InvoiceItem'
    field = 'originAmount'
    depth = 3
    max_depth = 2
    exclude_types = ['view']
    show_types = True
    
    if len(sys.argv) > 1:
        max_depth = int(sys.argv[1]) if sys.argv[1].isdigit() else 2
    if len(sys.argv) > 2:
        exclude_types_str = sys.argv[2]
        exclude_types = exclude_types_str.split(',') if exclude_types_str else []
    if len(sys.argv) > 3:
        show_types = sys.argv[3].lower() == 'true'
    
    rel_type = 0
    if 'view' not in exclude_types:
        rel_type = 0
    elif 'writeBack' in exclude_types and 'intra' not in exclude_types:
        rel_type = 2
    elif 'intra' in exclude_types and 'writeBack' not in exclude_types:
        rel_type = 1
    
    print(f"查询参数: objectType={object_type}, field={field}, depth={depth}, max_depth={max_depth}, exclude_types={exclude_types}, show_types={show_types}")
    
    data = fetch_impact_data(object_type, field, depth=depth, rel_type=rel_type, direction='downstream')
    create_excel(data, 'invoice_item_origin_amount_impact.xlsx', max_depth=max_depth, exclude_types=exclude_types, show_types=show_types)
